import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

#data_only=true로 해야 수식이 아닌 값으로 받아옴
load_wb = load_workbook("mymusic.xlsx",data_only=True)
load_ws = load_wb['Sheet1']

dates = ['20200301',	'20200302',	'20200303',	'20200304',	'20200305',	'20200306',	'20200307',	'20200308',	'20200309',	'20200310',	'20200311',	'20200312',	'20200313',	'20200314',	'20200315',	'20200316',	'20200317',	'20200318',	'20200319',	'20200320',	'20200321',	'20200322',	'20200323',	'20200324',	'20200325',	'20200326',	'20200327',	'20200328',	'20200329',	'20200330']
j=2
for date in dates:

    headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
    data = requests.get('https://www.genie.co.kr/chart/top200?ditc=D&ymd='+date+'&hh=23&rtm=N&pg=1',headers=headers)
    soup = BeautifulSoup(data.text, 'html.parser')
    songs = soup.select('#body-content > div.newest-list > div > table > tbody > tr')

    load_ws.cell(1,j,date)


    i=2
    for song in songs:
        rank = song.select_one('td.number').text[0:2].strip()
        title = song.select_one('td.info > a.title.ellipsis').text.strip()
        artist = song.select_one('td.info > a.artist.ellipsis').text

        load_ws.cell(i,j,title+':'+artist)
        i += 1

    j +=1

load_wb.save("mymusic.xlsx")

